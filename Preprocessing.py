import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys

# Redirect standard output to a file
sys.stdout = open('output.txt', 'w')

# Read the Excel file
data = pd.read_excel('Form Responses.xlsx', engine='openpyxl')

# Define all possible time slots for each day
time_slots = [
    "9:30AM-10:30AM", "10:30AM-11:30AM", "11:30AM-12:30PM",
    "12:30PM-2PM", "2PM-3:30PM", "3:30PM-5PM", "5:30PM-7:00PM", "7PM-8:30PM"
]

# Initialize the new DataFrame
columns = (
    ["First Name", "Last Name", "Email ID", "Roll Number","Contact Number","Department","Interview Happened"] +
    ["Amritansh", "Sara","Hridyansh"] + 
    [f"{day} {slot}" for day in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"] for slot in time_slots]
)
new_df = pd.DataFrame(columns=columns)

# Function to populate the slot availability
def create_slot_availability(row, day_column_name):
    slots = row.get(day_column_name, "").split(", ") if isinstance(row.get(day_column_name, ""), str) else []
    availability = {f"{day_column_name.split()[0]} {slot}": (1 if slot in slots else 0) for slot in time_slots}
    return availability

# Populate the DataFrame
processed_data = []
for index, row in data.iterrows():
    # Extract basic details
    first_name = row.get("First Name", "")
    last_name = row.get("Last Name", "")
    email_id = row.get("Email ID", "")
    roll_number = row.get("Roll Number", "")
    department = row.get("Department", "")
    contact = row.get("Contact Number", "")

    # Initialize a dictionary with personal details
    row_data = {
        "First Name": first_name,
        "Last Name": last_name,
        "Email ID": email_id,
        "Contact Number": contact,
        "Amritansh": "",
        "Sara": "",
        "Hridyansh": "",
        "Interview Happened": 0,
        "Roll Number": roll_number,
        "Department": department
    }

    # Add slot availability for each day
    for day in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]:
        row_data.update(create_slot_availability(row, day))
    
    # Add placeholder values for smpc1, smpc2, and smpc3
    row_data["Amritansh"] = ""
    row_data["Sara"] = ""
    row_data["Hridyansh"] = ""
    row_data["Interview Happened"] = 0
    processed_data.append(row_data)

# Convert the processed data into a DataFrame
new_df = pd.DataFrame(processed_data)

# Fill NaN values with 0 for slot availability columns
slot_columns = [col for col in new_df.columns if any(day in col for day in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"])]
new_df[slot_columns] = new_df[slot_columns].fillna(0)

# Save the DataFrame to an Excel file
new_df.to_excel('test_with_colors.xlsx', index=False, engine="openpyxl")

# Apply conditional formatting using openpyxl
wb = load_workbook('test_with_colors.xlsx')
ws = wb.active

# Define green fill for 1
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Apply the formatting
for row in ws.iter_rows(min_row=2, min_col=5, max_col=ws.max_column):
    for cell in row:
        if cell.value == 1:
            cell.fill = green_fill

# Save the workbook with conditional formatting applied
wb.save('test_with_colors.xlsx')

# Prepare the final DataFrame
columns_to_keep = (
    ["First Name", "Last Name", "Email ID", "Roll Number","Department","Contact Number", "Amritansh", "Sara","Hridyansh","Interview Happened"] +
    [col for col in new_df.columns if any(day in col for day in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"])]
)
df_final = new_df[columns_to_keep]

# Save the final DataFrame with formatting
df_final.to_excel('test_with_colors.xlsx', index=False, engine='openpyxl')

# Re-apply conditional formatting to the final file
wb = load_workbook('test_with_colors.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=2, min_col=5, max_col=ws.max_column):
    for cell in row:
        if cell.value == 1:
            cell.fill = green_fill

# Save the final styled workbook
wb.save('test_with_colors.xlsx')

# Close the file to reset standard output
sys.stdout.close()

# Reset standard output back to the terminal
sys.stdout = sys.__stdout__

print("The output has been saved to 'output.txt' and 'test_with_colors.xlsx'.")
